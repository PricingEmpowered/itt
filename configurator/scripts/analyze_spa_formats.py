from pathlib import Path
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb

FILES = [
    Path("/home/ubuntu/upload/CollinsCornelius_WorksheetV3.xlsb"),
    Path("/home/ubuntu/upload/SP651888-Extract.xlsx"),
    Path("/home/ubuntu/upload/SPA_ITEMS-PartsView.xlsx"),
]

def nonempty(values):
    return ["" if value is None else str(value).strip() for value in values]

def preview_rows(rows, max_rows=18, max_cols=30):
    for row_index, row in enumerate(rows[:max_rows], start=1):
        cleaned = nonempty(row[:max_cols])
        if any(cleaned):
            print(f"  Row {row_index}: {cleaned}")

def analyze_xlsx(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    print(f"\n=== {path.name} ===")
    print(f"Sheets: {workbook.sheetnames}")
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        populated = [row for row in rows if any(value is not None and str(value).strip() for value in row)]
        print(f"\nSheet: {sheet.title} | rows={len(rows)} | populated_rows={len(populated)} | max_col={sheet.max_column}")
        preview_rows(populated)

def analyze_xlsb(path: Path):
    print(f"\n=== {path.name} ===")
    with open_xlsb(path) as workbook:
        sheets = workbook.sheets
        print(f"Sheets: {sheets}")
        for sheet_name in sheets:
            with workbook.get_sheet(sheet_name) as sheet:
                rows = [[cell.v for cell in row] for row in sheet.rows()]
            populated = [row for row in rows if any(value is not None and str(value).strip() for value in row)]
            max_col = max((len(row) for row in rows), default=0)
            print(f"\nSheet: {sheet_name} | rows={len(rows)} | populated_rows={len(populated)} | max_col={max_col}")
            preview_rows(populated)

for file_path in FILES:
    if not file_path.exists():
        print(f"Missing: {file_path}")
        continue
    if file_path.suffix.lower() == ".xlsb":
        analyze_xlsb(file_path)
    else:
        analyze_xlsx(file_path)
